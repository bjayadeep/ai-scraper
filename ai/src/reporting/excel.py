import datetime
import logging
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import settings
from src.profiles import profile_value

logger = logging.getLogger(__name__)

def generate_styled_excel(jobs: List[Dict[str, Any]], profile: Any) -> str:
    """
    Generates a beautifully styled Excel file using pandas and openpyxl.
    Saves to history directory with naming CyberJobs_DDMMYYYY.xlsx.
    If file is already open/locked, appends a counter suffix.

    Returns:
        str: Absolute path of the generated Excel file.
    """
    # Create filename with current date: CyberJobs_DDMMYYYY.xlsx
    today_str = datetime.date.today().strftime("%d%m%Y")
    profile_slug = profile_value(profile, "slug", "jobs")
    profile_name = profile_value(profile, "name", "Jobs")
    base_filename = f"{profile_slug}_{today_str}"
    file_path = settings.HISTORY_DIR / f"{base_filename}.xlsx"

    # Handle locked file — append _v2, _v3 etc. if file is open in Excel
    counter = 2
    while file_path.exists():
        try:
            with file_path.open("r+b"):
                pass
            break  # File is writable — overwrite it
        except PermissionError:
            file_path = settings.HISTORY_DIR / f"{base_filename}_v{counter}.xlsx"
            counter += 1
    
    logger.info(f"Generating styled Excel at: {file_path}")
    
    # 1. Map jobs list to DataFrame
    data = []
    for idx, job in enumerate(jobs, 1):
        data.append({
            "S.No": idx,
            "Date Added": job.get("date_posted") or datetime.date.today().strftime("%Y-%m-%d"),
            "Company": job.get("company", ""),
            "Job Title": job.get("title", ""),
            "Location": job.get("location", ""),
            "Experience Required": job.get("experience_metadata", "Not Specified"),
            "Apply Link": job.get("apply_link", "")
        })
        
    df = pd.DataFrame(data)
    
    # Ensure folder exists
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 2. Write to Excel using openpyxl engine
    writer = pd.ExcelWriter(file_path, engine="openpyxl")
    sheet_name = profile_name[:31]
    df.to_excel(writer, index=False, startrow=3, sheet_name=sheet_name)
    
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Ensure grid lines are visible
    worksheet.views.sheetView[0].showGridLines = True
    
    # --- Styling Definition ---
    font_family = "Segoe UI"
    
    # Color palette
    brand_blue = "1F4E79"        # Deep Slate Blue for main headers
    accent_blue = "DDEBF7"       # Soft Blue for table headers background
    zebra_color = "F2F6FA"       # Alternating row background
    link_color = "0563C1"        # Classic link blue
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E79")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="595959")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    cell_font = Font(name=font_family, size=10)
    link_font = Font(name=font_family, size=10, color=link_color, underline="single")
    
    # Fills
    header_fill = PatternFill(start_color=brand_blue, end_color=brand_blue, fill_type="solid")
    zebra_fill = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=Side(border_style="medium", color="1F4E79"), bottom=Side(border_style="medium", color="1F4E79"))
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 3. Add Report Title Block
    worksheet["A1"] = f"🇺🇸 {profile_name.upper()} JOB LEADS (1-6 YRS EXP)"
    worksheet["A1"].font = title_font
    
    generated_time = datetime.datetime.now().strftime("%B %d, %Y - %I:%M %p")
    worksheet["A2"] = f"Report Generated: {generated_time} | Total Jobs: {len(jobs)}"
    worksheet["A2"].font = subtitle_font
    
    # Set height of title/subtitle rows
    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 18
    worksheet.row_dimensions[4].height = 28  # Header row height
    
    # 4. Format headers (Row 4 is where pandas headers end up because startrow=3 is 0-indexed offset + 1)
    # Excel rows are 1-indexed. startrow=3 means pandas headers are on Row 4.
    header_row = 4
    num_cols = len(df.columns)
    
    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_idx in [1, 2] else left_align
        cell.border = header_border
        
    # 5. Format Data Rows (Row 5 onwards)
    data_start_row = 5
    for row_idx in range(data_start_row, data_start_row + len(df)):
        worksheet.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        current_fill = zebra_fill if is_even else white_fill
        
        for col_idx in range(1, num_cols + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.fill = current_fill
            cell.border = cell_border
            
            # Alignments
            if col_idx in [1, 2]:  # S.No and Date
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            # Formatting for Apply Link column
            if col_idx == 7:  # Apply Link column (1-indexed)
                link_url = cell.value
                if link_url and str(link_url).startswith("http"):
                    cell.hyperlink = link_url   # Makes it clickable
                    cell.value = link_url        # Keep actual URL so history dedup works
                    cell.font = link_font        # Blue underline style
                    cell.alignment = left_align
                    
    # 6. Auto-fit column widths
    for col in worksheet.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            # Avoid title cell length messing up column A width
            if cell.row < 4:
                continue
            val_str = str(cell.value or '')
            if cell.hyperlink:
                val_str = "Click to Apply"
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Give column standard padding
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 10)
        
    writer.close()
    logger.info(f"Successfully generated and formatted Excel report: {file_path}")
    return str(file_path.resolve())
